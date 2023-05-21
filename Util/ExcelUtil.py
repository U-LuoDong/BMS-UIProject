from openpyxl import load_workbook

from Conf.VarConfig import testData


class ExcelUtil:
    def __init__(self):
        self.workbook = None
        self.sheet = None

    def load_workbook(self, filename):
        # 加载名为filename的文件
        # load_workbook(path)：加载指定路径的excel文件
        try:
            self.workbook = load_workbook(filename)
        except Exception as e:
            print(e)

    def get_sheet_name(self, filename):
        # 通过sheetname获取sheet页
        self.sheet = self.workbook[filename]

    def row_numbers(self):
        # 返回当前页的行数
        # ws.max_row：工作表最大行数
        return self.sheet.max_row

    def col_numbers(self):
        # 返回当前页的列数
        # ws.max_column：工作表最大列数
        return self.sheet.max_column

    def row_values(self, row):
        # 获取某一行的所有值，返回一个列表
        columns = self.sheet.max_column
        row_data = []
        for i in range(1, columns + 1):
            cell_value = self.sheet.cell(row=row, column=i).value
            row_data.append(cell_value)
        return row_data

    def cell_name(self, row, column):
        # 获取某一单元格的值
        return self.sheet.cell(row=row, column=column).value


# 调试
if __name__ == '__main__':
    filename = (testData + "/BMSTestCase.xlsx")
    test_data1 = ExcelUtil()
    test_data1.load_workbook(filename)
    test_data1.get_sheet_name("Login")
    a = test_data1.row_values(1)
    b = test_data1.row_values(2)
    test_data1s = dict(zip(a, b))
    print(test_data1s)
    print(test_data1s["UserName"])
    print(test_data1s["PassWord"])



    # print(wk.cell_name(2, 5))
